import openpyxl

from openpyxl.utils import get_column_letter,column_index_from_string
print(get_column_letter(1))        # Translate column 1 to a letter.
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(get_column_letter(sheet.max_column))  # Prints the letter of the largest column in the sheet.

print(column_index_from_string('A'))        # Get A's number.
print(column_index_from_string('AA'))        # Get AA's number.

