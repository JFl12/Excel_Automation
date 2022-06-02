import openpyxl

wb = openpyxl.Workbook() #It starts with one sheet.

wb.create_sheet()   #Add a new sheet.
print(wb.sheetnames)

wb.create_sheet(index=0,title='First Sheet')    #Create a new sheet at index 0.
print(wb.sheetnames)

wb.create_sheet(index=2,title='Middle Sheet')
print(wb.sheetnames)

del wb['Middle Sheet']  # this deletes the 'Middle Sheet'
del wb['Sheet1']        #  this deletes 'Sheet1'
print(wb.sheetnames)

# DONT FORGET TO CALL THE SAVE() MEHTHOD TO SAVE THE CHANGES AFTER ADDING SHEETS TO OR REMOVING SHEETS FROM THE WORKBOOK

#wb.save('example_copy.xlsx')    #Save the workbook to a new file  ( not the original file)