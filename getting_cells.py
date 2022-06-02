import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']    # Get a sheet from the workbook.
sheet['A1']             # Get a cell from the sheet.
print(sheet['A1'].value) # Get the value from the cell.

c = sheet['B1'] # Get another cell from the sheet.
print(c.value)  # Print the cell from the sheet.

print(f'Row {c.row}, Column {c.column} is {c.value}')  # Get the row, column, and value from the cell.
print(f'Cell {c.coordinate} is {c.value}')
print(sheet['C1'].value)

#***************************************************

print(sheet.cell(row=1,column=2))
print(sheet.cell(row=1,column=2).value)
#************************************************************

for i in range(1,8,2):  #Go through every other row:
    print(i,sheet.cell(row=i,column=2).value)  # only print column 2

print(sheet.max_row)
print(sheet.max_column)

for i in range(1,sheet.max_row):  #Go through every row:
   print(i,sheet.cell(row=i,column=2).value)  # only prints column 2
#***********************************************************************

