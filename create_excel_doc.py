import openpyxl

wb = openpyxl.Workbook()  #Create a blank workbook

wb.sheetnames   #It starts with one sheet.
sheet = wb.active
print(sheet.title)

sheet.title ='Spam Spam Spam'  #change title.
print(wb.sheetnames)
wb.save('example_copy.xlsx')    #Save the workbook to a new file  ( not the original file)