#! python3
# updatedProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']


PRICE_UPDATES = {'Garlic': 3.07,        #The produce types and their updated prices
                 'Celery': 1.19,
                 'Lemon:': 1.27}

for rowNum in range(2,sheet.max_row):   # skip the  first row
    produceName = sheet.cell(row=rowNum,column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum,column=2).value = PRICE_UPDATES[produceName]
wb.save('updatedProduceSales.xlsx')
